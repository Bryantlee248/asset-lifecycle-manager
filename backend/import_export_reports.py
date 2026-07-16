"""批量导入导出与报表统计API模块（新台账模板v1.0）"""
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fastapi import UploadFile, File, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, or_
from datetime import date, timedelta, datetime, timezone
from typing import Optional
from urllib.parse import quote

from database import get_db, Asset, Procurement, Change, Fault, Warranty, Retirement, AssetInbound, AssetOutbound, record_stage_change
from constants import LIFECYCLE_STAGES
from config_cache import is_valid_enum, get_enum_values, FIELD_TO_GROUP_IMPORT


# ============ 下拉选项校验（A-08 / O8：运行期查 DB 缓存） ============
def validate_field(field_name: str, value: str) -> list[str]:
    """校验字段值是否在系统配置已启用的合法选项内，返回错误列表。

    - lifecycle_stage 保留硬编码常量（O4）。
    - 其余可映射字段经 config_cache.is_valid_enum 校验。
    - 无法映射的字段跳过（交由业务层处理）。
    """
    errors = []
    if not value:
        return errors
    if field_name == "lifecycle_stage":
        if value not in LIFECYCLE_STAGES:
            errors.append(f"字段'{field_name}'值'{value}'不在合法选项中: {LIFECYCLE_STAGES}")
        return errors
    if field_name in FIELD_TO_GROUP_IMPORT:
        if not is_valid_enum(FIELD_TO_GROUP_IMPORT[field_name], value):
            errors.append(f"字段'{field_name}'值'{value}'不在系统配置已启用的选项中")
    return errors


# ============ 批量导入 ============
async def import_assets_excel(file: UploadFile, db: Session):
    """批量导入资产台账数据（Excel文件）"""
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value).strip() if cell.value else "")

    expected_cols = {
        "资产编号": "asset_code", "资产分类": "asset_category", "品牌": "brand",
        "型号": "model", "SN号": "sn", "机房": "room",
        "机柜": "cabinet", "U位": "u_position",
        "生命周期阶段": "lifecycle_stage", "入场日期": "entry_date",
        "责任人": "responsible_person", "维保状态": "warranty_status",
        "维保到期日": "warranty_expire_date", "备注": "remarks",
        "设备名称": "device_name", "项目名称": "project_name",
        "产权归属": "ownership", "所属部门": "department",
        "原值(元)": "original_value",
    }

    col_map = {}
    for i, h in enumerate(headers):
        if h in expected_cols:
            col_map[expected_cols[h]] = i

    if "asset_code" not in col_map or "asset_category" not in col_map:
        raise HTTPException(400, "Excel缺少必要列：资产编号、资产分类")

    success_count = 0
    skip_count = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[col_map.get("asset_code", 0)]:
            continue

        data = {}
        asset_date_fields = {"entry_date", "warranty_expire_date"}
        asset_numeric_fields = {"original_value"}
        for field, col_idx in col_map.items():
            val = row[col_idx] if col_idx < len(row) else None
            if val is not None:
                if isinstance(val, datetime):
                    data[field] = val.date()
                elif isinstance(val, date):
                    data[field] = val
                elif field in asset_date_fields:
                    try:
                        data[field] = datetime.strptime(str(val).strip(), "%Y-%m-%d").date()
                    except (ValueError, TypeError):
                        data[field] = str(val).strip()
                elif field in asset_numeric_fields:
                    try:
                        data[field] = float(val)
                    except (ValueError, TypeError):
                        data[field] = 0.0
                else:
                    data[field] = str(val).strip()

        row_errors = []
        for field, value in data.items():
            row_errors.extend(validate_field(field, value))

        code = data.get("asset_code", "")
        existing = db.query(Asset).filter(Asset.asset_code == code).first()
        if existing:
            skip_count += 1
            errors.append(f"行{row_idx}: 编号{code}已存在，跳过")
            continue

        if row_errors:
            errors.append(f"行{row_idx}: {'; '.join(row_errors)}")
            continue

        if "lifecycle_stage" not in data:
            data["lifecycle_stage"] = "规划"

        try:
            asset = Asset(**data)
            db.add(asset)
            success_count += 1
            if success_count % 100 == 0:
                db.commit()
        except Exception as e:
            db.rollback()
            errors.append(f"行{row_idx}: 写入失败 - {str(e)}")

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        errors.append(f"批量提交失败 - {str(e)}")

    return {
        "success": success_count,
        "skipped": skip_count,
        "errors": errors[:50],
        "total_rows": ws.max_row - 1
    }


async def import_subtable_excel(file: UploadFile, table_type: str, db: Session):
    """批量导入分表数据"""
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    model_map = {
        "procurement": (Procurement, {
            "资产编号": "asset_code", "采购申请编号": "request_no", "供应商": "vendor",
            "设备名称": "device_name", "配置参数摘要": "config_summary",
            "申请日期": "request_date", "申请人": "applicant",
            "审批状态": "approval_status", "数量": "quantity", "单价": "unit_price",
            "总价": "total_price", "备注": "remarks"
        }),
        "change": (Change, {
            "资产编号": "asset_code", "变更类型": "change_type", "工单编号": "work_order_no",
            "变更内容": "change_content", "原配置": "old_config", "新配置": "new_config",
            "变更原因": "change_reason", "审批人": "approver", "执行人": "executor",
            "执行日期": "execute_date", "完成状态": "completion_status", "备注": "remarks"
        }),
        "fault": (Fault, {
            "资产编号": "asset_code", "故障等级": "fault_level", "故障单号": "fault_no",
            "故障现象": "fault_description", "故障日期": "fault_date", "维修人": "repair_person",
            "处理方式": "handle_method", "配件更换": "parts_replaced", "根因分类": "root_cause",
            "恢复日期": "recovery_date", "停机时长": "downtime_hours", "维修费用": "repair_cost",
            "是否复发": "is_recurring", "备注": "remarks"
        }),
        "warranty": (Warranty, {
            "资产编号": "asset_code", "维保单号": "warranty_no", "维保类型": "warranty_type",
            "维保供应商": "warranty_vendor", "合同编号": "contract_no", "覆盖范围": "coverage",
            "维保起始日": "start_date", "维保到期日": "end_date", "续保决策": "renewal_decision",
            "决策人": "decision_person", "决策日期": "decision_date",
            "续保合同号": "renewal_contract_no", "续保起始日": "renewal_start_date",
            "续保到期日": "renewal_end_date", "维保费用": "cost", "备注": "remarks"
        }),
        "retirement": (Retirement, {
            "资产编号": "asset_code", "报废原因": "retire_reason", "报废类别": "retire_category",
            "申请单号": "application_no", "审批人": "approver", "审批日期": "approval_date",
            "下架日期": "uninstall_date", "下架人": "uninstall_person",
            "数据清除确认": "data_cleared", "数据清除人": "data_clear_person",
            "处置方式": "disposal_method", "残值回收": "residual_value", "备注": "remarks"
        }),
        "inbound": (AssetInbound, {
            "资产编号": "asset_code", "移入单号": "inbound_no", "接收类型": "receive_type",
            "产权归属": "ownership", "产权方公司": "owner_company", "项目名称": "project_name",
            "项目序号": "project_no", "资产分类": "asset_category", "品牌": "brand",
            "型号": "model", "SN序列号": "sn", "配置参数摘要": "config_summary",
            "采购合同编号": "purchase_contract_no", "采购总价": "purchase_total_price",
            "移入日期": "inbound_date", "接收人": "receiver",
            "验收结果": "inspection_result", "存放位置": "storage_location", "备注": "remarks"
        }),
        "outbound": (AssetOutbound, {
            "资产编号": "asset_code", "移出单号": "outbound_no", "移出原因": "outbound_reason",
            "移出类别": "outbound_category", "去向/目的地": "destination",
            "移出日期": "outbound_date", "接收方联系人": "receiver_contact",
            "接收方联系电话": "receiver_phone", "操作人": "operator",
            "审批人": "approver", "备注": "remarks"
        }),
    }

    if table_type not in model_map:
        raise HTTPException(400, f"不支持的表类型: {table_type}")

    Model, col_mapping = model_map[table_type]

    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    col_map = {}
    for i, h in enumerate(headers):
        if h in col_mapping:
            col_map[col_mapping[h]] = i

    if "asset_code" not in col_map:
        raise HTTPException(400, "Excel缺少必要列：资产编号")

    success_count = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[col_map.get("asset_code", 0)]:
            continue

        data = {}
        # 日期字段列表（需将字符串转为Python date对象）
        date_fields = {
            "entry_date", "warranty_expire_date",
            "request_date", "fault_date", "recovery_date", "execute_date",
            "start_date", "end_date", "decision_date", "renewal_start_date", "renewal_end_date",
            "approval_date", "uninstall_date",
            "inbound_date", "outbound_date",
        }
        for field, col_idx in col_map.items():
            val = row[col_idx] if col_idx < len(row) else None
            if val is not None:
                if isinstance(val, datetime):
                    data[field] = val.date()
                elif isinstance(val, date):
                    data[field] = val
                elif field in date_fields:
                    # 字符串日期转为Python date对象
                    try:
                        data[field] = datetime.strptime(str(val).strip(), "%Y-%m-%d").date()
                    except (ValueError, TypeError):
                        data[field] = str(val).strip()
                elif field in ("quantity", "unit_price", "total_price", "downtime_hours", "cost", "residual_value", "repair_cost", "purchase_total_price"):
                    try:
                        data[field] = float(val)
                        if field == "quantity":
                            data[field] = int(data[field])
                    except (ValueError, TypeError):
                        data[field] = None
                elif field == "is_recurring":
                    data[field] = str(val).strip() in ("是", "True", "true", "1", "复发")
                else:
                    data[field] = str(val).strip()

        row_errors = []
        for field, value in data.items():
            if isinstance(value, str):
                row_errors.extend(validate_field(field, value))

        if row_errors:
            errors.append(f"行{row_idx}: {'; '.join(row_errors)}")
            continue

        try:
            item = Model(**data)
            db.add(item)
            if table_type == "fault" and data.get("fault_level") in ("P1", "P2-严重"):
                asset = db.query(Asset).filter(Asset.asset_code == data.get("asset_code")).first()
                if asset and asset.lifecycle_stage in ["运行", "上架"]:
                    orig_stage = asset.lifecycle_stage
                    asset.lifecycle_stage = "维修"
                    record_stage_change(
                        db, asset.asset_code, orig_stage, "维修",
                        datetime.now(timezone.utc), operator="system_import",
                        reason="Excel导入故障降级", is_backfill=False
                    )
            success_count += 1
            if success_count % 100 == 0:
                db.commit()
        except Exception as e:
            db.rollback()
            errors.append(f"行{row_idx}: 写入失败 - {str(e)}")

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        errors.append(f"批量提交失败 - {str(e)}")

    return {"success": success_count, "errors": errors[:50], "total_rows": ws.max_row - 1}


# ============ 批量导出 ============
def export_assets_excel(
    db: Session,
    category: Optional[str] = None,
    stage: Optional[str] = None,
    warranty_status: Optional[str] = None,
    search: Optional[str] = None
) -> StreamingResponse:
    """导出资产台账为Excel文件"""
    query = db.query(Asset)
    if category:
        query = query.filter(Asset.asset_category == category)
    if stage:
        query = query.filter(Asset.lifecycle_stage == stage)
    if warranty_status:
        query = query.filter(Asset.warranty_status == warranty_status)
    if search:
        search_pattern = f"%{search}%"
        query = query.filter(or_(
            Asset.asset_code.ilike(search_pattern),
            Asset.brand.ilike(search_pattern),
            Asset.model.ilike(search_pattern),
            Asset.sn.ilike(search_pattern),
            Asset.room.ilike(search_pattern),
            Asset.device_name.ilike(search_pattern),
            Asset.responsible_person.ilike(search_pattern),
        ))

    assets = query.order_by(Asset.asset_code).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "资产台账主索引"

    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0052D9", end_color="0052D9", fill_type="solid")
    cell_font = Font(name="微软雅黑", size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    red_fill = PatternFill(start_color="FDE8E8", end_color="FDE8E8", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF7E8", end_color="FFF7E8", fill_type="solid")

    headers_list = ["资产编号", "资产分类", "品牌", "型号", "SN号", "机房",
                    "机柜", "U位", "生命周期阶段", "入场日期", "责任人", "维保状态",
                    "维保到期日", "设备名称", "项目名称", "产权归属", "原值(元)", "备注"]

    for col, h in enumerate(headers_list, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    today = date.today()
    for row_idx, a in enumerate(assets, 2):
        values = [
            a.asset_code, a.asset_category, a.brand, a.model, a.sn, a.room,
            a.cabinet, a.u_position, a.lifecycle_stage, str(a.entry_date) if a.entry_date else "",
            a.responsible_person, a.warranty_status,
            str(a.warranty_expire_date) if a.warranty_expire_date else "",
            a.device_name or "", a.project_name or "", a.ownership or "",
            round(float(a.original_value or 0), 2), a.remarks or ""
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = cell_font
            cell.alignment = center_align
            cell.border = thin_border

        if a.warranty_expire_date and a.lifecycle_stage in ("上架", "运行", "维修"):
            if a.warranty_expire_date < today:
                for col in range(1, len(headers_list) + 1):
                    ws.cell(row=row_idx, column=col).fill = red_fill
            elif a.warranty_expire_date < today + timedelta(days=30):
                ws.cell(row=row_idx, column=13).fill = yellow_fill

    col_widths = [16, 12, 10, 16, 16, 14, 14, 12, 14, 12, 10, 10, 12, 14, 14, 10, 14, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"资产台账导出_{today.strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"}
    )


def export_subtable_excel(db: Session, table_type: str) -> StreamingResponse:
    """导出分表为Excel"""
    table_config = {
        "procurement": (Procurement, "采购入库", [
            ("资产编号", "asset_code"), ("采购申请编号", "request_no"), ("供应商", "vendor"),
            ("设备名称", "device_name"), ("配置参数摘要", "config_summary"),
            ("申请日期", "request_date"), ("申请人", "applicant"),
            ("审批状态", "approval_status"), ("数量", "quantity"), ("单价", "unit_price"),
            ("总价", "total_price"), ("备注", "remarks")
        ]),
        "change": (Change, "变更迁移", [
            ("资产编号", "asset_code"), ("变更类型", "change_type"), ("工单编号", "work_order_no"),
            ("变更内容", "change_content"), ("原配置", "old_config"), ("新配置", "new_config"),
            ("变更原因", "change_reason"), ("审批人", "approver"), ("执行人", "executor"),
            ("执行日期", "execute_date"), ("完成状态", "completion_status"), ("备注", "remarks")
        ]),
        "fault": (Fault, "故障维修", [
            ("资产编号", "asset_code"), ("故障等级", "fault_level"), ("故障单号", "fault_no"),
            ("故障现象", "fault_description"), ("故障日期", "fault_date"),
            ("维修人", "repair_person"), ("处理方式", "handle_method"),
            ("配件更换", "parts_replaced"), ("根因分类", "root_cause"),
            ("恢复日期", "recovery_date"), ("停机时长", "downtime_hours"),
            ("维修费用", "repair_cost"), ("是否复发", "is_recurring"), ("备注", "remarks")
        ]),
        "warranty": (Warranty, "维保续保", [
            ("资产编号", "asset_code"), ("维保单号", "warranty_no"), ("维保类型", "warranty_type"),
            ("维保供应商", "warranty_vendor"), ("合同编号", "contract_no"), ("覆盖范围", "coverage"),
            ("维保起始日", "start_date"), ("维保到期日", "end_date"), ("续保决策", "renewal_decision"),
            ("决策人", "decision_person"), ("决策日期", "decision_date"),
            ("续保合同号", "renewal_contract_no"), ("续保起始日", "renewal_start_date"),
            ("续保到期日", "renewal_end_date"), ("维保费用", "cost"), ("备注", "remarks")
        ]),
        "retirement": (Retirement, "退役报废", [
            ("资产编号", "asset_code"), ("报废原因", "retire_reason"), ("报废类别", "retire_category"),
            ("申请单号", "application_no"), ("审批人", "approver"), ("审批日期", "approval_date"),
            ("下架日期", "uninstall_date"), ("下架人", "uninstall_person"),
            ("数据清除确认", "data_cleared"), ("数据清除人", "data_clear_person"),
            ("处置方式", "disposal_method"), ("残值回收", "residual_value"), ("备注", "remarks")
        ]),
        "inbound": (AssetInbound, "资产移入", [
            ("资产编号", "asset_code"), ("移入单号", "inbound_no"), ("接收类型", "receive_type"),
            ("产权归属", "ownership"), ("产权方公司", "owner_company"), ("项目名称", "project_name"),
            ("项目序号", "project_no"), ("资产分类", "asset_category"), ("品牌", "brand"),
            ("型号", "model"), ("SN序列号", "sn"), ("配置参数摘要", "config_summary"),
            ("采购合同编号", "purchase_contract_no"), ("采购总价", "purchase_total_price"),
            ("移入日期", "inbound_date"), ("接收人", "receiver"),
            ("验收结果", "inspection_result"), ("存放位置", "storage_location"), ("备注", "remarks")
        ]),
        "outbound": (AssetOutbound, "资产移出", [
            ("资产编号", "asset_code"), ("移出单号", "outbound_no"), ("移出原因", "outbound_reason"),
            ("移出类别", "outbound_category"), ("去向/目的地", "destination"),
            ("移出日期", "outbound_date"), ("接收方联系人", "receiver_contact"),
            ("接收方联系电话", "receiver_phone"), ("操作人", "operator"),
            ("审批人", "approver"), ("备注", "remarks")
        ]),
    }

    if table_type not in table_config:
        raise HTTPException(400, f"不支持的表类型: {table_type}")

    Model, sheet_name, col_defs = table_config[table_type]
    items = db.query(Model).order_by(Model.id.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0052D9", end_color="0052D9", fill_type="solid")
    cell_font = Font(name="微软雅黑", size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    for col, (header, _) in enumerate(col_defs, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for row_idx, item in enumerate(items, 2):
        for col, (_, field) in enumerate(col_defs, 1):
            val = getattr(item, field, None)
            if isinstance(val, date):
                val = str(val)
            elif isinstance(val, bool):
                val = "是" if val else "否"
            cell = ws.cell(row=row_idx, column=col, value=val or "")
            cell.font = cell_font
            cell.alignment = center_align
            cell.border = thin_border

            if table_type == "fault" and field == "fault_level" and val in ("P1", "P2-严重"):
                cell.fill = PatternFill(start_color="FDE8E8", end_color="FDE8E8", fill_type="solid")
                cell.font = Font(name="微软雅黑", size=10, bold=True, color="E34D59")

    for col in range(1, len(col_defs) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 14

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{sheet_name}导出_{date.today().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"}
    )


def download_import_template(table_type: str) -> StreamingResponse:
    """下载导入模板（含示例数据和下拉选项提示）"""
    wb = openpyxl.Workbook()
    ws = wb.active

    template_config = {
        "assets": ("资产台账导入模板", [
            ("资产编号*", "资产分类*", "品牌", "型号", "SN号", "机房",
             "机柜", "U位", "生命周期阶段", "入场日期", "责任人", "维保状态",
             "维保到期日", "设备名称", "项目名称", "产权归属", "原值(元)", "备注"),
            [("DC-CL-SRV-001", "服务器", "Dell", "R740", "SN001", "5-4机房",
              "R-03", "15-16U", "规划", "2026-01-15", "张三", "在保",
              "2029-01-15", "核心交换机", "数据中心项目", "自有", 50000, "示例数据")]
        ]),
        "procurement": ("采购入库导入模板", [
            ("资产编号*", "采购申请编号", "供应商", "设备名称", "配置参数摘要",
             "申请日期", "申请人", "审批状态", "数量", "单价", "总价", "备注"),
            [("DC-CL-SRV-001", "PR-2026-001", "戴尔科技", "PowerEdge R740",
             "2*E5-2680v4/64G/2*600G SSD", "2026-01-10", "李四", "审批中",
             1, 50000, 50000, "示例")]
        ]),
        "change": ("变更迁移导入模板", [
            ("资产编号*", "变更类型", "工单编号", "变更内容", "原配置", "新配置",
             "变更原因", "审批人", "执行人", "执行日期", "完成状态", "备注"),
            [("DC-CL-SVR-001", "配置变更", "WO-2026-001", "内存扩容",
             "64G", "128G", "业务增长需要", "运维主管", "王五",
             "2026-06-01", "已完成", "示例")]
        ]),
        "fault": ("故障维修导入模板", [
            ("资产编号*", "故障等级", "故障单号", "故障现象", "故障日期", "维修人", "处理方式",
             "配件更换", "根因分类", "恢复日期", "停机时长", "维修费用", "是否复发", "备注"),
            [("DC-CL-SVR-001", "P3", "FLT-2026-001", "端口故障", "2026-06-01", "王五", "现场修复",
             "无", "硬件故障", "2026-06-02", 4, 500, "否", "示例")]
        ]),
        "warranty": ("维保续保导入模板", [
            ("资产编号*", "维保单号", "维保类型", "维保供应商", "合同编号", "覆盖范围",
             "维保起始日", "维保到期日", "续保决策",
             "决策人", "决策日期", "续保合同号", "续保起始日", "续保到期日", "维保费用", "备注"),
            [("DC-CL-SVR-001", "WB-2026-001", "整机维保", "戴尔科技", "CT-2026-001", "整机维保",
             "2026-01-15", "2029-01-15", "续保",
             "运维主管", "2026-01-01", "WB-2029-001", "2029-01-15", "2032-01-15", 5000, "示例")]
        ]),
        "retirement": ("退役报废导入模板", [
            ("资产编号*", "报废原因", "报废类别", "申请单号", "审批人", "审批日期",
             "下架日期", "下架人", "数据清除确认", "数据清除人", "处置方式", "残值回收", "备注"),
            [("DC-CL-SVR-001", "设备老化", "报废", "RF-2026-001", "运维主管", "2026-06-01",
             "2026-06-05", "王五", "已清除", "安全负责人", "回收商处理", 500, "示例")]
        ]),
        "inbound": ("资产移入导入模板", [
            ("资产编号*", "移入单号", "接收类型", "产权归属", "产权方公司", "项目名称",
             "项目序号", "资产分类", "品牌", "型号", "SN序列号", "配置参数摘要",
             "采购合同编号", "采购总价", "移入日期", "接收人", "验收结果", "存放位置", "备注"),
            [("DC-CL-SVR-001", "IN-2026-001", "采购入库", "自有", "长乐数据中心", "数据中心扩容",
             "PRJ-001", "服务器", "Dell", "PowerEdge R740", "SN-IN001", "2*E5-2680v4/64G",
             "CT-2026-001", 50000, "2026-01-15", "张三", "合格", "5-4机房 R-03", "示例")]
        ]),
        "outbound": ("资产移出导入模板", [
            ("资产编号*", "移出单号", "移出原因", "移出类别", "去向/目的地",
             "移出日期", "接收方联系人", "接收方联系电话", "操作人", "审批人", "备注"),
            [("DC-CL-SVR-001", "OUT-2026-001", "设备老化报废", "报废", "回收商",
             "2026-06-05", "李经理", "13800138000", "王五", "运维主管", "示例")]
        ]),
    }

    if table_type not in template_config:
        raise HTTPException(400, f"不支持的模板类型: {table_type}")

    sheet_name, (headers, examples) = template_config[table_type]
    ws.title = sheet_name

    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0052D9", end_color="0052D9", fill_type="solid")
    tip_font = Font(name="微软雅黑", size=9, color="8C8E96")
    cell_font = Font(name="微软雅黑", size=10)
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    center_align = Alignment(horizontal="center", vertical="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    tip_map_assets = {
        2: "/".join(get_enum_values("category")),
        7: "/".join(LIFECYCLE_STAGES),
        10: "/".join(get_enum_values("warranty_status")),
    }
    tip_map_proc = {10: "/".join(get_enum_values("procurement_approval_status"))}
    tip_map_fault = {2: "/".join(get_enum_values("fault_level")), 6: "/".join(get_enum_values("handle_method")), 8: "/".join(get_enum_values("root_cause"))}
    tip_map_warranty = {6: "/".join(get_enum_values("renewal_decision"))}
    tip_map_retirement = {3: "/".join(get_enum_values("retire_category")), 9: "/".join(get_enum_values("data_clear_option"))}
    tip_map_change = {2: "/".join(get_enum_values("change_type")), 13: "/".join(get_enum_values("completion_status"))}
    tip_map_inbound = {3: "/".join(get_enum_values("receive_type")), 4: "/".join(get_enum_values("ownership_type")), 8: "/".join(get_enum_values("category")), 17: "/".join(get_enum_values("inbound_inspection_result"))}
    tip_map_outbound = {4: "/".join(get_enum_values("outbound_category"))}

    all_tip_maps = {
        "assets": tip_map_assets, "procurement": tip_map_proc,
        "fault": tip_map_fault, "warranty": tip_map_warranty,
        "retirement": tip_map_retirement, "change": tip_map_change,
        "inbound": tip_map_inbound, "outbound": tip_map_outbound,
    }
    current_tips = all_tip_maps.get(table_type, {})

    for col, tip in current_tips.items():
        cell = ws.cell(row=2, column=col, value=f"可选: {tip}")
        cell.font = tip_font
        cell.fill = PatternFill(start_color="F3F5F8", end_color="F3F5F8", fill_type="solid")

    for col, val in enumerate(examples[0], 1):
        cell = ws.cell(row=3, column=col, value=val)
        cell.font = cell_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")

    for row in range(4, 24):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col, value="")
            cell.border = thin_border

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{sheet_name}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"}
    )


# ============ 报表统计 ============
def get_comprehensive_report(db: Session):
    """综合报表：资产概览+分类分布+阶段分布+维保状态+故障概览"""
    today = date.today()
    total = db.query(Asset).count()

    by_category = dict(db.query(Asset.asset_category, func.count(Asset.id)).group_by(Asset.asset_category).all())
    by_stage = dict(db.query(Asset.lifecycle_stage, func.count(Asset.id)).group_by(Asset.lifecycle_stage).all())
    by_warranty = dict(db.query(Asset.warranty_status, func.count(Asset.id)).group_by(Asset.warranty_status).all())

    warranty_expired = db.query(Asset).filter(
        Asset.warranty_expire_date < today,
        Asset.lifecycle_stage.in_(["上架", "运行", "维修"])
    ).all()
    warranty_expiring = db.query(Asset).filter(
        Asset.warranty_expire_date.between(today, today + timedelta(days=90)),
        Asset.lifecycle_stage.in_(["上架", "运行", "维修"])
    ).all()

    total_faults = db.query(Fault).count()
    by_fault_level = dict(db.query(Fault.fault_level, func.count(Fault.id)).group_by(Fault.fault_level).all())
    by_root_cause = dict(db.query(Fault.root_cause, func.count(Fault.id)).group_by(Fault.root_cause).all())
    unresolved = db.query(Fault).filter(Fault.recovery_date == None).count()
    avg_downtime = db.query(func.avg(Fault.downtime_hours)).filter(Fault.downtime_hours != None).scalar() or 0

    age_buckets = {"0-1年": 0, "1-3年": 0, "3-5年": 0, "5-8年": 0, "8年以上": 0}
    for a in db.query(Asset).filter(Asset.entry_date != None).all():
        if a.entry_date:
            age_days = (today - a.entry_date).days
            age_years = age_days / 365.25
            if age_years < 1: age_buckets["0-1年"] += 1
            elif age_years < 3: age_buckets["1-3年"] += 1
            elif age_years < 5: age_buckets["3-5年"] += 1
            elif age_years < 8: age_buckets["5-8年"] += 1
            else: age_buckets["8年以上"] += 1

    by_change_type = dict(db.query(Change.change_type, func.count(Change.id)).group_by(Change.change_type).all())
    total_changes = db.query(Change).count()

    total_purchase_cost = db.query(func.sum(Procurement.total_price)).scalar() or 0

    return {
        "total_assets": total,
        "by_category": by_category,
        "by_stage": by_stage,
        "by_warranty": by_warranty,
        "warranty_expired_count": len(warranty_expired),
        "warranty_expired_list": [{"code": a.asset_code, "category": a.asset_category, "expire_date": str(a.warranty_expire_date), "brand": a.brand, "model": a.model} for a in warranty_expired[:20]],
        "warranty_expiring_count": len(warranty_expiring),
        "warranty_expiring_list": [{"code": a.asset_code, "category": a.asset_category, "expire_date": str(a.warranty_expire_date), "days_left": (a.warranty_expire_date - today).days} for a in warranty_expiring[:20]],
        "fault_summary": {
            "total": total_faults,
            "by_level": by_fault_level,
            "by_root_cause": by_root_cause,
            "unresolved": unresolved,
            "avg_downtime_hours": round(avg_downtime, 1)
        },
        "age_distribution": age_buckets,
        "change_summary": {"total": total_changes, "by_type": by_change_type},
        "total_purchase_cost": total_purchase_cost,
    }


def get_warranty_expiry_report(db: Session, days: int = 90):
    """维保到期报表：指定天数内即将到期 + 已过期"""
    today = date.today()
    cutoff = today + timedelta(days=days)

    expired = db.query(Asset).filter(
        Asset.warranty_expire_date < today,
        Asset.lifecycle_stage.in_(["上架", "运行", "维修"])
    ).all()

    expiring = db.query(Asset).filter(
        Asset.warranty_expire_date.between(today, cutoff),
        Asset.lifecycle_stage.in_(["上架", "运行", "维修"])
    ).all()

    return {
        "expired": [{"asset_code": a.asset_code, "category": a.asset_category, "brand": a.brand, "model": a.model, "room": a.room, "cabinet": a.cabinet, "responsible_person": a.responsible_person, "warranty_status": a.warranty_status, "warranty_expire_date": str(a.warranty_expire_date), "days_overdue": (today - a.warranty_expire_date).days} for a in expired],
        "expiring": [{"asset_code": a.asset_code, "category": a.asset_category, "brand": a.brand, "model": a.model, "room": a.room, "cabinet": a.cabinet, "responsible_person": a.responsible_person, "warranty_status": a.warranty_status, "warranty_expire_date": str(a.warranty_expire_date), "days_left": (a.warranty_expire_date - today).days} for a in expiring],
        "expired_count": len(expired),
        "expiring_count": len(expiring),
    }


def get_fault_analysis_report(db: Session, start_date: Optional[str] = None, end_date: Optional[str] = None):
    """故障分析报表：故障频率/根因/P级分布/设备故障排行"""
    query = db.query(Fault)
    if start_date:
        query = query.filter(Fault.fault_date >= start_date)
    if end_date:
        query = query.filter(Fault.fault_date <= end_date)

    faults = query.all()

    by_level = {}
    by_root_cause = {}
    by_asset = {}
    by_handle_method = {}
    total_downtime = 0.0
    recurring_count = 0

    for f in faults:
        by_level[f.fault_level] = by_level.get(f.fault_level, 0) + 1
        rc = f.root_cause or "未分类"
        by_root_cause[rc] = by_root_cause.get(rc, 0) + 1
        by_asset[f.asset_code] = by_asset.get(f.asset_code, 0) + 1
        hm = f.handle_method or "未分类"
        by_handle_method[hm] = by_handle_method.get(hm, 0) + 1
        if f.downtime_hours:
            total_downtime += f.downtime_hours
        if f.is_recurring:
            recurring_count += 1

    top_fault_assets = sorted(by_asset.items(), key=lambda x: x[1], reverse=True)[:10]

    unresolved_by_level = {}
    for f in db.query(Fault).filter(Fault.recovery_date == None).all():
        unresolved_by_level[f.fault_level] = unresolved_by_level.get(f.fault_level, 0) + 1

    return {
        "total_faults": len(faults),
        "by_level": by_level,
        "by_root_cause": by_root_cause,
        "by_handle_method": by_handle_method,
        "total_downtime_hours": round(total_downtime, 1),
        "avg_downtime_hours": round(total_downtime / max(len(faults), 1), 1),
        "recurring_count": recurring_count,
        "top_fault_assets": [{"asset_code": code, "fault_count": count} for code, count in top_fault_assets],
        "unresolved_by_level": unresolved_by_level,
        "unresolved_total": sum(unresolved_by_level.values()),
    }


def get_change_frequency_report(db: Session):
    """变更频率报表"""
    by_type = dict(db.query(Change.change_type, func.count(Change.id)).group_by(Change.change_type).all())
    by_month = {}
    for c in db.query(Change).filter(Change.execute_date != None).all():
        month_key = c.execute_date.strftime("%Y-%m")
        by_month[month_key] = by_month.get(month_key, 0) + 1

    by_month = dict(sorted(by_month.items()))

    by_asset = {}
    for c in db.query(Change).all():
        by_asset[c.asset_code] = by_asset.get(c.asset_code, 0) + 1

    top_change_assets = sorted(by_asset.items(), key=lambda x: x[1], reverse=True)[:10]

    total = db.query(Change).count()
    completed = db.query(Change).filter(Change.completion_status == "已完成").count()

    return {
        "total": total,
        "completed": completed,
        "completion_rate": round(completed / max(total, 1) * 100, 1),
        "by_type": by_type,
        "by_month": by_month,
        "top_change_assets": [{"asset_code": code, "change_count": count} for code, count in top_change_assets],
    }
