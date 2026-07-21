"""统计看板聚合模块 — 报表统计模块（新台账模板v1.0）

提供看板所需的 6 个聚合纯函数 + APIRouter(prefix="/api/stats") + Excel 导出函数。
所有接口复用 require_permission("reports:view")，不新增权限项。
聚合函数风格对齐 import_export_reports.py（纯函数(db) -> dict），便于测试与复用。

路由清单（前缀 /api/stats）：
  GET /overview              KPI 概览
  GET /stage-distribution    生命周期阶段分布（7 阶段全覆盖）
  GET /category-composition  分类/型号构成与原值
  GET /reliability           可靠性与故障分析（MTBF / 各阶段故障率 / TopN）
  GET /warranty-buckets      维保到期预警分桶
  GET /aggregate             自定义字段聚合（白名单 + count/original_value）
  GET /export                看板 Excel 导出（多 sheet）
"""
import io
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import date, datetime, timedelta
from calendar import monthrange
from urllib.parse import quote

from database import get_db, Asset, Fault, Retirement, AssetStageLog
from constants import (
    LIFECYCLE_STAGES,
    ACTIVE_STAGES,
    AGGREGATE_FIELD_WHITELIST,
)
from config_cache import get_category_code, build_aggregate_fields_cache, is_valid_aggregate_field, get_enabled_aggregate_fields
from auth import require_permission


# ============ 路由定义 ============
stats_router = APIRouter(prefix="/api/stats", tags=["统计看板"])


# ============ 1. KPI 概览 ============
def get_overview(db: Session) -> dict:
    """看板顶部 4 张 KPI 卡片数据：总资产 / 总原值 / 故障总数 / 30天内即将到期"""
    today = date.today()
    total_assets = db.query(Asset).count()
    total_original_value = float(
        db.query(func.coalesce(func.sum(Asset.original_value), 0.0)).scalar() or 0
    )
    total_faults = db.query(Fault).count()
    warranty_expiring_soon = db.query(Asset).filter(
        Asset.warranty_expire_date != None,
        Asset.warranty_expire_date.between(today, today + timedelta(days=30)),
        Asset.lifecycle_stage.in_(ACTIVE_STAGES),
    ).count()
    return {
        "total_assets": total_assets,
        "total_original_value": round(total_original_value, 2),
        "total_faults": total_faults,
        "warranty_expiring_soon": warranty_expiring_soon,
    }


# ============ 2. 生命周期阶段分布 ============
def get_stage_distribution(db: Session) -> dict:
    """7 阶段全覆盖，ratio = count / total（Σratio ≈ 1）"""
    total = db.query(Asset).count()
    counts = dict(
        db.query(Asset.lifecycle_stage, func.count(Asset.id))
        .group_by(Asset.lifecycle_stage).all()
    )
    stages = []
    for stage in LIFECYCLE_STAGES:
        c = counts.get(stage, 0)
        ratio = round(c / total, 4) if total else 0
        stages.append({"stage": stage, "count": c, "ratio": ratio})
    return {"stages": stages}


# ============ 3. 分类/型号构成与原值 ============
def get_category_composition(db: Session, category: Optional[str] = None, include_code: bool = False, stage: Optional[str] = None) -> dict:
    """按 asset_category 与 model 汇总台数与原值；支持按 category 过滤型号；可选返回分类码；S-12 联动支持 stage 过滤"""
    # 分类维度
    cat_q = db.query(
        Asset.asset_category,
        func.count(Asset.id),
        func.coalesce(func.sum(Asset.original_value), 0.0),
    )
    if stage:
        cat_q = cat_q.filter(Asset.lifecycle_stage == stage)
    cat_rows = cat_q.group_by(Asset.asset_category).all()
    by_category = []
    for cat, cnt, ov in cat_rows:
        item = {
            "category": cat or "未填写",
            "count": cnt,
            "original_value": round(float(ov or 0), 2),
        }
        if include_code:
            item["category_code"] = get_category_code(cat) if cat else "OTH"
        by_category.append(item)

    # 型号维度（可选按分类过滤 / 按阶段过滤）
    model_q = db.query(
        Asset.asset_category,
        Asset.model,
        func.count(Asset.id),
        func.coalesce(func.sum(Asset.original_value), 0.0),
    )
    if category:
        model_q = model_q.filter(Asset.asset_category == category)
    if stage:
        model_q = model_q.filter(Asset.lifecycle_stage == stage)
    model_rows = model_q.group_by(Asset.asset_category, Asset.model).all()
    by_model = [
        {
            "category": cat or "未填写",
            "model": model or "未填写",
            "count": cnt,
            "original_value": round(float(ov or 0), 2),
        }
        for cat, model, cnt, ov in model_rows
    ]
    return {"by_category": by_category, "by_model": by_model}


# ============ 4. 可靠性与故障分析 ============
def get_reliability(db: Session, top_n: int = 10, stage: Optional[str] = None) -> dict:
    """故障总数、系统级 MTBF、各阶段故障率、TopN 故障资产排行；S-12 联动支持 stage 过滤"""
    today = date.today()

    # 系统级 MTBF：Σ(各资产运营天数) / 全部故障次数
    # 运营天数 = min(退役下架日 uninstall_date, 今天) − entry_date，仅计 entry_date 非空资产
    assets_with_entry_q = db.query(Asset).filter(Asset.entry_date != None)
    if stage:
        assets_with_entry_q = assets_with_entry_q.filter(Asset.lifecycle_stage == stage)
    assets_with_entry = assets_with_entry_q.all()
    total_operational_days = 0
    for a in assets_with_entry:
        ret = db.query(Retirement).filter(Retirement.asset_code == a.asset_code).first()
        end_date = ret.uninstall_date if (ret and ret.uninstall_date) else today
        days = (end_date - a.entry_date).days
        total_operational_days += max(0, days)  # 防止脏数据(入场日>今天)导致负值

    total_faults = db.query(Fault).count()
    mtbf_days = round(total_operational_days / total_faults, 1) if total_faults else 0

    # 各阶段故障率（提供 stage 时仅返回该阶段）
    stages_to_compute = [stage] if stage else LIFECYCLE_STAGES
    by_stage_failure_rate = []
    for s in stages_to_compute:
        asset_count = db.query(Asset).filter(Asset.lifecycle_stage == s).count()
        fault_count = 0
        if asset_count > 0:
            codes = [a.asset_code for a in db.query(Asset).filter(Asset.lifecycle_stage == s).all()]
            fault_count = db.query(Fault).filter(Fault.asset_code.in_(codes)).count() if codes else 0
        rate = round(fault_count / asset_count, 4) if asset_count else 0
        by_stage_failure_rate.append({
            "stage": s,
            "asset_count": asset_count,
            "fault_count": fault_count,
            "rate": rate,
        })

    # TopN 故障资产（提供 stage 时仅统计当前处于该阶段的资产）
    top_q = db.query(Fault.asset_code, func.count(Fault.id)).group_by(Fault.asset_code)
    if stage:
        codes_in_stage = [a.asset_code for a in db.query(Asset).filter(Asset.lifecycle_stage == stage).all()]
        top_q = top_q.filter(Fault.asset_code.in_(codes_in_stage)) if codes_in_stage else top_q.filter(False)
    top_rows = top_q.order_by(func.count(Fault.id).desc()).limit(top_n).all()
    top_fault_assets = [{"asset_code": code, "fault_count": cnt} for code, cnt in top_rows]

    return {
        "total_faults": total_faults,
        "mtbf_days": mtbf_days,
        "by_stage_failure_rate": by_stage_failure_rate,
        "top_fault_assets": top_fault_assets,
    }


# ============ 5. 维保到期预警分桶 ============
def get_warranty_buckets(db: Session, stage: Optional[str] = None) -> dict:
    """按到期时间分桶（已过期/30天内/60天内/90天内/90天以上），限 ACTIVE_STAGES；含即将到期清单；S-12 联动支持 stage 过滤"""
    today = date.today()
    if stage:
        # 指定阶段时直接按该阶段过滤（不再受 ACTIVE_STAGES 约束，以支持联动下钻到任意阶段）
        assets = db.query(Asset).filter(
            Asset.lifecycle_stage == stage,
            Asset.warranty_expire_date != None,
        ).all()
    else:
        assets = db.query(Asset).filter(
            Asset.lifecycle_stage.in_(ACTIVE_STAGES),
            Asset.warranty_expire_date != None,
        ).all()

    buckets = {"expired": 0, "within_30": 0, "within_60": 0, "within_90": 0, "over_90": 0}
    expiring_list = []
    for a in assets:
        d = a.warranty_expire_date
        days_left = (d - today).days
        if d < today:
            buckets["expired"] += 1
        elif days_left <= 30:
            buckets["within_30"] += 1
        elif days_left <= 60:
            buckets["within_60"] += 1
        elif days_left <= 90:
            buckets["within_90"] += 1
        else:
            buckets["over_90"] += 1
        # 即将到期清单：90天内（含已过期）
        if days_left <= 90:
            expiring_list.append({
                "asset_code": a.asset_code,
                "category": a.asset_category,
                "brand": a.brand,
                "model": a.model,
                "warranty_expire_date": str(d),
                "days_left": days_left,
                "responsible_person": a.responsible_person,
            })
    expiring_list.sort(key=lambda x: x["days_left"])
    return {"buckets": buckets, "expiring_list": expiring_list}


# ============ 6. 自定义字段聚合 ============
def get_aggregate(db: Session, field: str, metric: str = "count", stage: Optional[str] = None) -> dict:
    """对白名单字段 GROUP BY，返回各取值的计数与原值汇总；非法字段/指标返回 400；S-12 联动支持 stage 过滤。

    单一集成入口：读 aggregate_whitelist 表（经进程内缓存），仅允许启用中的字段；
    表空（极端未 seed）时回退常量 AGGREGATE_FIELD_WHITELIST（design §2 保留回退）。
    """
    valid = is_valid_aggregate_field(field)
    if valid is None:
        build_aggregate_fields_cache(db)
        valid = is_valid_aggregate_field(field)
    if not valid:
        raise HTTPException(status_code=400, detail=f"非法聚合字段: {field}（不在白名单内）")
    if metric not in ("count", "original_value"):
        raise HTTPException(status_code=400, detail=f"非法聚合指标: {metric}（仅支持 count/original_value）")

    col = getattr(Asset, field)
    q = db.query(
        col,
        func.count(Asset.id),
        func.coalesce(func.sum(Asset.original_value), 0.0),
    )
    if stage:
        q = q.filter(Asset.lifecycle_stage == stage)
    rows_db = q.group_by(col).all()

    rows = []
    for value, cnt, ov in rows_db:
        rows.append({
            "value": value if value not in (None, "") else "未填写",
            "count": cnt,
            "original_value": round(float(ov or 0), 2),
        })

    # 按当前指标降序
    sort_key = "original_value" if metric == "original_value" else "count"
    rows.sort(key=lambda r: r[sort_key], reverse=True)
    return {"field": field, "metric": metric, "rows": rows}


# ============ 7. 阶段分布趋势（S-16：按月末快照） ============
def _resolve_stage_at(db: Session, end: date) -> dict:
    """返回 end 月末时各资产的阶段（基于 AssetStageLog 快照：取 changed_at<=end 的最后一条记录的 to_stage）"""
    end_dt = datetime(end.year, end.month, end.day, 23, 59, 59)
    logs = db.query(AssetStageLog).filter(AssetStageLog.changed_at <= end_dt).all()
    stage_of: dict = {}
    for lg in logs:
        prev = stage_of.get(lg.asset_code)
        if prev is None or lg.changed_at > prev[1]:
            stage_of[lg.asset_code] = (lg.to_stage, lg.changed_at)
    return {code: v[0] for code, v in stage_of.items()}


def get_stage_trend(db: Session, months: int = 12) -> dict:
    """按月末快照统计各阶段资产数。

    口径：每月末 = 各资产「changed_at ≤ 该月末 的最后一条日志的 to_stage」；
    由于每台资产至少有「出生」锚点，故恒被计入，Σcounts = 总资产数。
    """
    months = max(1, min(60, int(months)))
    today = date.today()
    # 构造最近 months 个月的月末日期列表（升序）
    ends = []
    y, m = today.year, today.month
    for _ in range(months):
        ends.append(date(y, m, monthrange(y, m)[1]))
        m -= 1
        if m == 0:
            m = 12
            y -= 1
    ends.sort()

    stages = LIFECYCLE_STAGES
    matrix = []

    all_logs = db.query(AssetStageLog).all()
    logs_by_asset: dict = {}
    for lg in all_logs:
        logs_by_asset.setdefault(lg.asset_code, []).append(lg)

    for end in ends:
        end_dt = datetime(end.year, end.month, end.day, 23, 59, 59)
        counts = {s: 0 for s in stages}
        total = 0
        for code, logs in logs_by_asset.items():
            cand = [l for l in logs if l.changed_at <= end_dt]
            if not cand:
                continue
            last = max(cand, key=lambda l: l.changed_at)
            counts[last.to_stage] = counts.get(last.to_stage, 0) + 1
            total += 1
        matrix.append({"month": end.strftime("%Y-%m"), "counts": counts, "total": total})

    is_backfill = any(l.is_backfill for l in all_logs)
    return {
        "months": [e.strftime("%Y-%m") for e in ends],
        "stages": stages,
        "matrix": matrix,
        "is_backfill": is_backfill,
    }


# ============ 8. 时间范围对比（S-17：环比/同比） ============
def _month_snapshot(db: Session, ym: str, metric: str) -> dict:
    """返回 ym（YYYY-MM）月末快照；metric==stage 返回各阶段计数，否则返回单值指标"""
    y, m = map(int, ym.split("-"))
    end = date(y, m, monthrange(y, m)[1])
    stage_of_asset = _resolve_stage_at(db, end)

    if metric == "stage":
        counts = {s: 0 for s in LIFECYCLE_STAGES}
        for s in stage_of_asset.values():
            counts[s] += 1
        return {"month": ym, "snapshot": counts}

    assets = db.query(Asset).filter(Asset.asset_code.in_(list(stage_of_asset.keys()))).all() if stage_of_asset else []
    if metric == "total_assets":
        val = len(assets)
    elif metric == "active_assets":
        val = sum(1 for a in assets if a.lifecycle_stage in ACTIVE_STAGES)
    elif metric == "original_value":
        val = round(float(sum((a.original_value or 0) for a in assets)), 2)
    elif metric == "fault_count":
        start = date(y, m, 1)
        val = db.query(Fault).filter(Fault.fault_date >= start, Fault.fault_date <= end).count()
    else:
        val = 0
    return {"month": ym, "snapshot": val}


def get_stage_compare(db: Session, range_a: str, range_b: str, metric: str = "stage") -> dict:
    """对比两个月份窗口的月末快照，返回 a/b 快照、Δ、Δ%、compare_type（环比/同比/自定义）"""
    a = _month_snapshot(db, range_a, metric)
    b = _month_snapshot(db, range_b, metric)

    if metric == "stage":
        delta = {s: b["snapshot"][s] - a["snapshot"][s] for s in LIFECYCLE_STAGES}
        delta_pct = {
            s: round((b["snapshot"][s] - a["snapshot"][s]) / a["snapshot"][s] * 100, 2)
            if a["snapshot"][s] else None
            for s in LIFECYCLE_STAGES
        }
    else:
        av, bv = a["snapshot"], b["snapshot"]
        delta = bv - av
        delta_pct = round((bv - av) / av * 100, 2) if av else None

    # 由 range_a / range_b 推导对比类型
    ay, am = map(int, range_a.split("-"))
    by, bm = map(int, range_b.split("-"))
    if by == ay + 1 and bm == am:
        compare_type = "同比"
    elif (by * 12 + bm) == (ay * 12 + am) + 1:
        compare_type = "环比"
    else:
        compare_type = "自定义"

    return {
        "metric": metric,
        "a": a,
        "b": b,
        "delta": delta,
        "delta_pct": delta_pct,
        "compare_type": compare_type,
    }


# ============ 路由实现 ============
@stats_router.get("/overview")
def api_overview(db: Session = Depends(get_db), _: object = Depends(require_permission("reports:view"))):
    """KPI 概览：总资产/总原值/故障总数/30天内即将到期"""
    return get_overview(db)


@stats_router.get("/stage-distribution")
def api_stage_distribution(db: Session = Depends(get_db), _: object = Depends(require_permission("reports:view"))):
    """生命周期阶段分布（7 阶段全覆盖）"""
    return get_stage_distribution(db)


@stats_router.get("/category-composition")
def api_category_composition(
    category: Optional[str] = Query(None, description="按分类过滤型号"),
    include_code: bool = Query(False, description="是否返回分类码 category_code"),
    stage: Optional[str] = Query(None, description="S-12 联动：按阶段过滤"),
    db: Session = Depends(get_db),
    _: object = Depends(require_permission("reports:view")),
):
    """分类/型号构成与原值；可选 ?category= & ?include_code=true & ?stage="""
    return get_category_composition(db, category, include_code, stage)


@stats_router.get("/reliability")
def api_reliability(
    top_n: int = Query(10, ge=1, le=100, description="TopN 故障资产数量"),
    stage: Optional[str] = Query(None, description="S-12 联动：按阶段过滤"),
    db: Session = Depends(get_db),
    _: object = Depends(require_permission("reports:view")),
):
    """可靠性与故障分析：MTBF / 各阶段故障率 / TopN 故障资产；可选 ?stage= 联动下钻"""
    return get_reliability(db, top_n, stage)


@stats_router.get("/warranty-buckets")
def api_warranty_buckets(
    stage: Optional[str] = Query(None, description="S-12 联动：按阶段过滤"),
    db: Session = Depends(get_db),
    _: object = Depends(require_permission("reports:view")),
):
    """维保到期预警分桶（限 ACTIVE_STAGES）+ 即将到期清单；可选 ?stage= 联动下钻"""
    return get_warranty_buckets(db, stage)


@stats_router.get("/aggregate")
def api_aggregate(
    field: Optional[str] = Query(None, description="白名单内的聚合字段"),
    metric: str = Query("count", description="count | original_value"),
    stage: Optional[str] = Query(None, description="S-12 联动：按阶段过滤"),
    db: Session = Depends(get_db),
    _: object = Depends(require_permission("reports:view")),
):
    """自定义字段聚合：?field=<字段>&metric=count|original_value；非法字段/指标返回 400；可选 ?stage= 联动下钻"""
    if not field:
        raise HTTPException(status_code=400, detail="缺少聚合字段 field 参数")
    return get_aggregate(db, field, metric, stage)


@stats_router.get("/aggregate-fields")
def api_aggregate_fields(db: Session = Depends(get_db), _: object = Depends(require_permission("reports:view"))):
    """T-11：返回当前 enabled=true 的聚合维度 [{value,label}]，供统计页下拉实时同步。

    与 config:manage 分离，复用 reports:view 只读权限；极端未 seed 时回退常量。"""
    fields = get_enabled_aggregate_fields()
    if fields is None:
        build_aggregate_fields_cache(db)
        fields = get_enabled_aggregate_fields() or []
    if not fields:
        # 极端回退：seed 未跑时退化为常量，保证统计页不白屏（design §3.6）
        fields = [{"value": f, "label": f} for f in AGGREGATE_FIELD_WHITELIST]
    return fields


@stats_router.get("/stage-trend")
def api_stage_trend(
    months: int = Query(12, ge=1, le=60, description="最近月份数（1~60）"),
    db: Session = Depends(get_db),
    _: object = Depends(require_permission("reports:view")),
):
    """阶段分布趋势：返回最近 months 个月末各阶段资产数矩阵；含 is_backfill 标注"""
    return get_stage_trend(db, months)


@stats_router.get("/compare")
def api_compare(
    range_a: str = Query(..., description="对比窗口A（YYYY-MM）"),
    range_b: str = Query(..., description="对比窗口B（YYYY-MM）"),
    metric: str = Query("stage", description="stage | total_assets | active_assets | original_value | fault_count"),
    db: Session = Depends(get_db),
    _: object = Depends(require_permission("reports:view")),
):
    """时间范围对比：两窗口月末快照 + Δ + Δ%；compare_type 由后端推导（环比/同比/自定义）"""
    if metric not in ("stage", "total_assets", "active_assets", "original_value", "fault_count"):
        raise HTTPException(status_code=400, detail=f"非法对比指标: {metric}")
    for r in (range_a, range_b):
        try:
            y, m = map(int, r.split("-"))
            if not (1 <= m <= 12):
                raise ValueError
        except Exception:
            raise HTTPException(status_code=400, detail=f"非法月份格式: {r}（应为 YYYY-MM）")
    return get_stage_compare(db, range_a, range_b, metric)


# ============ 7. 看板 Excel 导出（T7） ============
def export_stats_excel(db: Session) -> StreamingResponse:
    """将当前看板视图（KPI + 各维度汇总）导出为多 sheet Excel"""
    overview = get_overview(db)
    stage = get_stage_distribution(db)
    comp = get_category_composition(db, include_code=True)
    rel = get_reliability(db, top_n=20)
    wbuckets = get_warranty_buckets(db)

    # 样式
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0052D9", end_color="0052D9", fill_type="solid")
    cell_font = Font(name="微软雅黑", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))

    def write_sheet(ws, title: str, headers: list, rows: list) -> None:
        ws.title = title
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin
        for ri, row in enumerate(rows, 2):
            for ci, val in enumerate(row, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = cell_font
                cell.alignment = center
                cell.border = thin
        for ci in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 16

    wb = openpyxl.Workbook()

    # Sheet1: KPI 汇总
    ws = wb.active
    write_sheet(ws, "KPI汇总", ["指标", "数值"], [
        ["总资产数", overview["total_assets"]],
        ["总原值(元)", overview["total_original_value"]],
        ["故障总数", overview["total_faults"]],
        ["30天内即将到期", overview["warranty_expiring_soon"]],
    ])

    # Sheet2: 阶段分布
    ws2 = wb.create_sheet()
    write_sheet(ws2, "阶段分布", ["阶段", "数量", "占比"],
                [[s["stage"], s["count"], s["ratio"]] for s in stage["stages"]])

    # Sheet3: 分类构成
    ws3 = wb.create_sheet()
    write_sheet(ws3, "分类构成", ["分类", "分类码", "数量", "原值(元)"],
                [[c["category"],
                  c.get("category_code", ""),
                  c["count"],
                  c["original_value"]] for c in comp["by_category"]])

    # Sheet4: 型号构成
    ws4 = wb.create_sheet()
    write_sheet(ws4, "型号构成", ["分类", "型号", "数量", "原值(元)"],
                [[m["category"], m["model"], m["count"], m["original_value"]] for m in comp["by_model"]])

    # Sheet5: 可靠性分析 + Top故障资产
    ws5 = wb.create_sheet()
    write_sheet(ws5, "可靠性分析", ["阶段", "资产数", "故障数", "故障率"],
                [[s["stage"], s["asset_count"], s["fault_count"], s["rate"]]
                 for s in rel["by_stage_failure_rate"]])
    top_start = len(rel["by_stage_failure_rate"]) + 3
    ws5.cell(row=top_start, column=1, value="Top故障资产（按故障次数）").font = Font(bold=True)
    ws5.cell(row=top_start + 1, column=1, value="资产编号")
    ws5.cell(row=top_start + 1, column=2, value="故障次数")
    for i, t in enumerate(rel["top_fault_assets"], start=top_start + 2):
        ws5.cell(row=i, column=1, value=t["asset_code"])
        ws5.cell(row=i, column=2, value=t["fault_count"])

    # Sheet6: 维保分桶
    ws6 = wb.create_sheet()
    b = wbuckets["buckets"]
    write_sheet(ws6, "维保分桶", ["分桶", "数量"], [
        ["已过期", b["expired"]],
        ["30天内", b["within_30"]],
        ["60天内", b["within_60"]],
        ["90天内", b["within_90"]],
        ["90天以上", b["over_90"]],
    ])

    # Sheet7: 即将到期清单
    ws7 = wb.create_sheet()
    write_sheet(ws7, "即将到期清单",
                ["资产编号", "分类", "品牌", "型号", "到期日", "剩余天数", "责任人"],
                [[e["asset_code"], e.get("category", ""), e.get("brand", ""), e.get("model", ""),
                  e["warranty_expire_date"], e["days_left"], e.get("responsible_person", "")]
                 for e in wbuckets["expiring_list"]])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"统计看板导出_{date.today().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


@stats_router.get("/export")
def api_export(db: Session = Depends(get_db), _: object = Depends(require_permission("reports:view"))):
    """导出当前看板视图为 Excel（多 sheet）"""
    return export_stats_excel(db)
